from typing import List, Optional, Dict, Any
from datetime import datetime
import uuid
import csv
import json
import io
from fastapi import HTTPException, UploadFile
from backend.models.sheet_models import SheetCreate, SheetResponse
from backend.models.transaction_models import TransactionCreate
from backend.utils.supabase_client import supabase
from backend.utils.logger import logger
from backend.utils.date_utils import DateUtils

class SheetService:
    """
    Service for managing financial sheets and importing transactions from various sources.
    """

    def create_sheet(self, sheet_data: SheetCreate, user_id: str) -> SheetResponse:
        """
        Create a new financial sheet.
        """
        try:
            sheet_id = str(uuid.uuid4())
            
            new_sheet = {
                "id": sheet_id,
                "name": sheet_data.name,
                "client_id": sheet_data.client_id,
                "financial_year": sheet_data.financial_year,
                "created_by": user_id,
                "created_at": datetime.utcnow().isoformat(),
                "updated_at": datetime.utcnow().isoformat()
            }
            
            data = supabase.table("sheets").insert(new_sheet).execute()
            
            if not data.data:
                raise HTTPException(status_code=500, detail="Failed to create sheet")
            
            return SheetResponse(**data.data[0])
            
        except Exception as e:
            logger.error(f"Error creating sheet: {e}")
            raise HTTPException(status_code=400, detail=str(e))

    def get_sheet(self, sheet_id: str) -> SheetResponse:
        """
        Retrieve a specific sheet by ID.
        """
        try:
            data = supabase.table("sheets").select("*").eq("id", sheet_id).is_("deleted_at", "null").execute()
            
            if not data.data:
                raise HTTPException(status_code=404, detail="Sheet not found")
            
            return SheetResponse(**data.data[0])
            
        except Exception as e:
            raise HTTPException(status_code=400, detail=str(e))

    def list_sheets(self, client_id: str) -> List[SheetResponse]:
        """
        List all sheets for a client.
        """
        try:
            data = supabase.table("sheets").select("*").eq("client_id", client_id).is_("deleted_at", "null").execute()
            
            return [SheetResponse(**sheet) for sheet in data.data]
            
        except Exception as e:
            raise HTTPException(status_code=400, detail=str(e))

    def delete_sheet(self, sheet_id: str) -> dict:
        """
        Soft delete a sheet.
        """
        try:
            data = supabase.table("sheets").update({
                "deleted_at": datetime.utcnow().isoformat()
            }).eq("id", sheet_id).execute()
            
            if not data.data:
                raise HTTPException(status_code=404, detail="Sheet not found")
            
            return {"success": True, "message": "Sheet deleted successfully"}
            
        except Exception as e:
            raise HTTPException(status_code=400, detail=str(e))

    def restore_sheet(self, sheet_id: str) -> SheetResponse:
        """
        Restore a soft-deleted sheet.
        """
        try:
            data = supabase.table("sheets").update({
                "deleted_at": None
            }).eq("id", sheet_id).execute()
            
            if not data.data:
                raise HTTPException(status_code=404, detail="Sheet not found")
            
            return SheetResponse(**data.data[0])
            
        except Exception as e:
            raise HTTPException(status_code=400, detail=str(e))

    async def import_csv(self, file: UploadFile, sheet_id: str) -> dict:
        """
        Import transactions from CSV file.
        """
        try:
            content = await file.read()
            csv_data = content.decode('utf-8')
            reader = csv.DictReader(io.StringIO(csv_data))
            
            transactions = []
            for row in reader:
                # Normalize keys to lower case
                row_lower = {k.lower(): v for k, v in row.items()}
                
                transaction = {
                    "id": str(uuid.uuid4()),
                    "sheet_id": sheet_id,
                    "date": DateUtils.parse_date(row_lower.get("date", "")),
                    "description": row_lower.get("description", "") or row_lower.get("narration", ""),
                    "amount": float(row_lower.get("amount", 0)),
                    "type": row_lower.get("type", "debit").lower(),
                    "ledger": row_lower.get("ledger", "Uncategorized"),
                    "created_at": datetime.utcnow().isoformat(),
                    "updated_at": datetime.utcnow().isoformat()
                }
                transactions.append(transaction)
            
            if transactions:
                # Use bulk insert logic
                supabase.table("transactions").insert(transactions).execute()
            
            return {"success": True, "count": len(transactions)}
            
        except Exception as e:
            logger.error(f"CSV import failed: {e}")
            raise HTTPException(status_code=400, detail=f"CSV import failed: {str(e)}")

    async def import_json(self, file: UploadFile, sheet_id: str) -> dict:
        """
        Import transactions from JSON file.
        """
        try:
            content = await file.read()
            json_data = json.loads(content)
            
            transactions = []
            for item in json_data:
                transaction = {
                    "id": str(uuid.uuid4()),
                    "sheet_id": sheet_id,
                    "date": DateUtils.parse_date(item.get("date", "")),
                    "description": item.get("description", ""),
                    "amount": float(item.get("amount", 0)),
                    "type": item.get("type", "debit"),
                    "ledger": item.get("ledger", "Uncategorized"),
                    "created_at": datetime.utcnow().isoformat(),
                    "updated_at": datetime.utcnow().isoformat()
                }
                transactions.append(transaction)
            
            if transactions:
                supabase.table("transactions").insert(transactions).execute()
            
            return {"success": True, "count": len(transactions)}
            
        except Exception as e:
            logger.error(f"JSON import failed: {e}")
            raise HTTPException(status_code=400, detail=f"JSON import failed: {str(e)}")

    async def import_excel(self, file: UploadFile, sheet_id: str) -> dict:
        """
        Import transactions from Excel file (.xlsx).
        """
        try:
            import openpyxl
            content = await file.read()
            workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            sheet = workbook.active
            
            transactions = []
            # Get headers from the first row
            headers = [str(cell.value).lower() if cell.value else f"col_{i}" for i, cell in enumerate(sheet[1])]
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not any(row): continue # Skip empty rows
                
                row_dict = dict(zip(headers, row))
                
                # Map common column names
                date_val = row_dict.get("date") or row_dict.get("transaction date")
                desc_val = row_dict.get("description") or row_dict.get("narration") or row_dict.get("particulars")
                amount_val = row_dict.get("amount") or row_dict.get("debit") or row_dict.get("credit")
                type_val = row_dict.get("type") or ("credit" if row_dict.get("credit") else "debit")
                
                transaction = {
                    "id": str(uuid.uuid4()),
                    "sheet_id": sheet_id,
                    "date": DateUtils.parse_date(str(date_val)) if date_val else None,
                    "description": str(desc_val) if desc_val else "",
                    "amount": float(amount_val) if amount_val else 0.0,
                    "type": str(type_val).lower(),
                    "ledger": str(row_dict.get("ledger", "Uncategorized")),
                    "created_at": datetime.utcnow().isoformat(),
                    "updated_at": datetime.utcnow().isoformat()
                }
                transactions.append(transaction)
            
            if transactions:
                supabase.table("transactions").insert(transactions).execute()
            
            return {"success": True, "count": len(transactions)}
            
        except ImportError:
             raise HTTPException(status_code=500, detail="openpyxl library not installed")
        except Exception as e:
            logger.error(f"Excel import failed: {e}")
            raise HTTPException(status_code=400, detail=f"Excel import failed: {str(e)}")

    async def import_zoho(self, file: UploadFile, sheet_id: str) -> dict:
        """
        Import transactions from Zoho Books export (CSV/XLS).
        """
        # Zoho Books logic is similar to generic CSV/Excel but with specific column mapping
        # For now, we reuse the generic importers which have fuzzy column matching
        filename = file.filename.lower()
        if filename.endswith(".csv"):
            return await self.import_csv(file, sheet_id)
        elif filename.endswith(".xlsx") or filename.endswith(".xls"):
            return await self.import_excel(file, sheet_id)
        else:
             raise HTTPException(status_code=400, detail="Unsupported file format for Zoho import")

    async def import_khatabook(self, file: UploadFile, sheet_id: str) -> dict:
        """
        Import transactions from Khatabook export.
        """
        # Khatabook logic is similar to generic CSV/Excel
        filename = file.filename.lower()
        if filename.endswith(".csv"):
            return await self.import_csv(file, sheet_id)
        elif filename.endswith(".xlsx") or filename.endswith(".xls"):
            return await self.import_excel(file, sheet_id)
        else:
             raise HTTPException(status_code=400, detail="Unsupported file format for Khatabook import")
